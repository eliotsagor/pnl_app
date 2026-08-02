"""Export database P&L to an Excel file matching the original layout."""
from datetime import date
from pathlib import Path
import calendar

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

import database


# Column layout: month -> (day_col_letter, value_col_letter)
# Jan A/B, Feb D/E, Mar G/H, Apr J/K, May M/N, Jun P/Q, Jul S/T, Aug V/W, Sep Y/Z, Oct AB/AC, Nov AE/AF, Dec AH/AI
MONTH_COLUMNS = {
    1:  ("A", "B"),
    2:  ("D", "E"),
    3:  ("G", "H"),
    4:  ("J", "K"),
    5:  ("M", "N"),
    6:  ("P", "Q"),
    7:  ("S", "T"),
    8:  ("V", "W"),
    9:  ("Y", "Z"),
    10: ("AB", "AC"),
    11: ("AE", "AF"),
    12: ("AH", "AI"),
}


def date_to_cell(d: date, value_col: bool = True) -> str:
    """Return the Excel cell coordinate for a given date.

    Day N goes to row N+1 (day 1 is row 2 since row 1 has month header).
    """
    day_col, val_col = MONTH_COLUMNS[d.month]
    col = val_col if value_col else day_col
    row = d.day + 1
    return f"{col}{row}"


def export_to_xlsx(template_path: Path, output_path: Path, year: int = 2026):
    """Write all database P&L data into a copy of the template workbook."""
    wb = load_workbook(template_path)
    label_to_sheet = database.get_label_to_sheet()
    all_data = database.get_all_pnl()

    # Group by date for formula generation
    by_date = {}
    for row in all_data:
        d = row["trade_date"]
        if isinstance(d, str):
            d = date.fromisoformat(d)
        by_date.setdefault(d, {})[row["sheet_name"]] = row["value"]

    # Map sheet name -> actual workbook sheet name (handles trailing spaces)
    sheet_lookup = {name.strip(): name for name in wb.sheetnames}

    bic_sheets = ["BIC $6 $4", "BIC $4", "BIC $3", "BIC 1 DTE", "BIC Standard", "BIC Re-Entry"]
    total_components = [
        "All BICs", "Price Action IC", "Verticals", "Auto JSP", "Wed Thu RIC",
        "Lottery", "MOC IC", "WUGA", "MT BF", "Umbrella", "PM Umbrella",
        "Vol Crush", "Lunch Vol C Umb", "LDOM",
    ]

    for d, values in by_date.items():
        if d.year != year:
            continue
        cell = date_to_cell(d, value_col=True)

        # Write direct strategy values
        for sheet_name, value in values.items():
            actual = sheet_lookup.get(sheet_name)
            if not actual or actual not in wb.sheetnames:
                continue
            wb[actual][cell] = value

        # Write All BICs formula if any BIC values exist for this day
        bic_refs = []
        for bic in bic_sheets:
            actual = sheet_lookup.get(bic)
            if actual and actual in wb.sheetnames:
                bic_refs.append(f"'{actual}'!{cell}")
        if bic_refs and "All BICs" in wb.sheetnames:
            wb["All BICs"][cell] = "=" + "+".join(bic_refs)

        # Write Total formula
        total_refs = []
        for comp in total_components:
            actual = sheet_lookup.get(comp)
            if actual and actual in wb.sheetnames:
                total_refs.append(f"{_quote_ref(actual)}!{cell}")
        # Include FOMC Meeting in Total only if there's a value that day
        if "FOMC Meeting" in values:
            actual = sheet_lookup.get("FOMC Meeting")
            if actual:
                total_refs.append(f"'{actual}'!{cell}")
        if total_refs and "Total" in wb.sheetnames:
            wb["Total"][cell] = "=" + "+".join(total_refs)

    wb.save(output_path)


def _quote_ref(sheet_name: str) -> str:
    """Quote sheet names that contain spaces or special chars."""
    if any(c in sheet_name for c in " $"):
        return f"'{sheet_name}'"
    return sheet_name
