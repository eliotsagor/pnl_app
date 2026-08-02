"""Seed the database from the master .xlsx workbook (one-off import)."""
from datetime import date
from pathlib import Path

import openpyxl

import database


WB_PATH = Path(r"C:\Users\Eliot\Downloads\2026- Daily PnL By Strategy.xlsx")
YEAR = 2026

MONTH_COLUMNS = {
    1:  ("A", "B"),
    2:  ("D", "E"),
    3:  ("G", "H"),
    4:  ("J", "K"),
    5:  ("M", "N"),
    6:  ("P", "Q"),
    7:  ("S", "T"),
    8:  ("V", "W"),
    9:  ("Y", "Z"),
    10: ("AB", "AC"),
    11: ("AE", "AF"),
    12: ("AH", "AI"),
}


def extract_from_workbook():
    wb = openpyxl.load_workbook(WB_PATH, data_only=True)
    sheet_lookup = {name.strip(): name for name in wb.sheetnames}

    strategies = [s["sheet_name"] for s in database.get_strategies()
                  if s["sheet_name"] != "All BICs"]

    data = {}
    for sheet in strategies:
        actual = sheet_lookup.get(sheet)
        if not actual:
            continue
        ws = wb[actual]
        for month, (day_col, val_col) in MONTH_COLUMNS.items():
            day_idx = openpyxl.utils.column_index_from_string(day_col)
            val_idx = openpyxl.utils.column_index_from_string(val_col)
            for r in range(2, 33):
                day = ws.cell(row=r, column=day_idx).value
                val = ws.cell(row=r, column=val_idx).value
                if val is None or not isinstance(val, (int, float)):
                    continue
                if not isinstance(day, (int, float)):
                    continue
                try:
                    trade_date = date(YEAR, month, int(day))
                except ValueError:
                    continue
                data.setdefault(trade_date, {})[sheet] = float(val)
    return data


def seed():
    database.reset_db()
    data = extract_from_workbook()
    for d in sorted(data):
        values = data[d]
        database.save_day(d, values)
        total = sum(values.values())
        print(f"Saved {d}: {len(values)} entries, total ${total:,.2f}")
    print(f"\nDone. {len(data)} days seeded.")


if __name__ == "__main__":
    seed()
